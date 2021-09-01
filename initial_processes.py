
"""
Created on Wed Nov 22 16:17:18 2017

@author: Alfredo Gonzalez-Sulser, University of Edinburgh
email: agonzal2@staffmail.ed.ac.uk
"""
import glob
from numpy import *
import pandas as pd
from scipy import spatial
from scipy.signal import decimate
from itertools import combinations
import parameters
import matplotlib.pyplot as plt
prm = parameters.Parameters()
from OpenEphys import *
import mne
import xlrd
import xlsxwriter
from openpyxl import load_workbook
import pathlib


def load_file(file):  #Opens text files.
    print(" Opening file " + file)

    data=loadtxt(file)
    print(len(data))
    return data




"The function below loads 16 channel headstage recordings individually, specify which one of a potential 4 to load"

def load_16channel_tetrode(folder_name, montage_name, downsampling, amp_filter):

    # headstage_number == 1:
    channels=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16]
    data=loadFolderToArray(folder_name, channels, chprefix = 'CH', dtype = float, session = '0', source = '100')
    data1 = decimate(data*1000, downsampling, axis = 0).astype(int16)
    # headstage_number == 2:
    channels=[17,18,19,20,21,22, 23,24,25,26,27,28,29,30,31,32]
    data=loadFolderToArray(folder_name, channels, chprefix = 'CH', dtype = float, session = '0', source = '100')
    data2 = decimate(data*1000, downsampling, axis = 0).astype(int16)
    # headstage_number == 3:
    channels=[33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48]
    data=loadFolderToArray(folder_name, channels, chprefix = 'CH', dtype = float, session = '0', source = '100')
    data3 = decimate(data*1000, downsampling, axis = 0).astype(int16)
    # headstage_number == 4:
    channels=[49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64]
    data=loadFolderToArray(folder_name, channels, chprefix = 'CH', dtype = float, session = '0', source = '100')
    data4 = decimate(data*1000, downsampling, axis = 0).astype(int16)
    
    # it returns values in nano Volts as integers.
    return data1, data2, data3, data4


def load_16channel_opto_individually(headstage_number):


    if headstage_number == 1:
        channels=[1,2,3,4,5,6,7,8,9,10,11,12,13,14,15,16]

        'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
        data=loadFolderToArray(prm.get_filepath(), channels, chprefix = 'CH', dtype = float, session = '0', source = '100')#######load file
        #data_adc=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'ADC', dtype = float, session = '0', source = '100')#######load file8

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        #data[:,16]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

    if headstage_number == 2:
        channels=[17,18,19,20,21,22, 23,24,25,26,27,28,29,30,31,32]

        'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
        data=loadFolderToArray(prm.get_filepath(), channels, chprefix = 'CH', dtype = float, session = '0', source = '100')#######load file
        #data_adc=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'ADC', dtype = float, session = '0', source = '100')#######load file8

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        #data[:,16]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

    if headstage_number == 3:
        channels=[33,34,35,36,37,38,39,40,41,42,43,44,45,46,47,48]

        'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
        data=loadFolderToArray(prm.get_filepath(), channels, chprefix = 'CH', dtype = float, session = '0', source = '100')#######load file
        #data_adc=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'ADC', dtype = float, session = '0', source = '100')#######load file8

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        #data[:,16]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

    if headstage_number == 4:
        channels=[49,50,51,52,53,54,55,56,57,58,59,60,61,62,63,64]

        'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
        data=loadFolderToArray(prm.get_filepath(), channels, chprefix = 'CH', dtype = float, session = '0', source = '100')#######load file
        #data_adc=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'ADC', dtype = float, session = '0', source = '100')#######load file8

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        #data[:,16]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.


    return data

def load_16channel_opto(headstage_number):

    'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
    data=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = '100')#######load file
    data_adc=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'ADC', dtype = float, session = '0', source = '100')#######load file8

    if headstage_number == 4:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,64]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.




    if headstage_number == 3:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,48]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.


    if headstage_number ==2:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,32]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

        n_channels=33


    if headstage_number ==1:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,16]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

    return data


"The function below loads the data in mne format, specify number of headstages"

def load_16_channel_opto_mne(headstage_number):


    'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
    data=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = '100')#######load file
    data_adc=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'ADC', dtype = float, session = '0', source = '100')#######load file8

    if headstage_number == 4:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,64]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

        datatp=data.transpose()#Array from openephys has to be transposed to match RawArray MNE function to create.
        del data
        del data_adc

        'Below I make the channel names and channel types, this should go in the parameteres file later.'

        n_channels=65

        channel_names=['hpc_mid_deep', 'hpc_mid_sup', 'hpc_ros_deep', 'hpc_ros_sup', 'pfc_deep',
                       'pfc_sup', 'cx1', 'cx2', 'hpc_contra_deep', 'hpc_contra_sup',
                       'EMG1', 'EMG2', 'cb_deep', 'cb_sup', 'hp_caudal_deep', 'hpc_caudal_sup',
                       'hpc_mid_d_2', 'hpc_mid_s_2', 'hpc_ros_d_2', 'hpc_ros_s_2', 'pfc_d_2',
                       'pfc_sup_2', 'cx1_2', 'cx2_2', 'hpc_ct_d_2', 'hpc_ct_s_2',
                       'EMG1_2', 'EMG2_2', 'cb_deep_2', 'cb_sup_2', 'hp_caud_d_2', 'hpc_caud_s_2',
                       'hpc_mid_d_3', 'hpc_mid_s_3', 'hpc_ros_d_3', 'hpc_ros_s_3', 'pfc_d_3',
                       'pfc_s_3', 'cx1_3', 'cx2_3', 'hpc_c_d_3', 'hpc_c_s_3',
                       'EMG1_3', 'EMG2_3', 'cb_deep_3', 'cb_sup_3', 'hp_c_d_3', 'hpc_c_s_3',
                       'hpc_mid_d_4', 'hpc_mid_s_4', 'hpc_ros_d_4', 'hpc_ros_s_4', 'pfc_d_4',
                       'pfc_s_4', 'cx1_4', 'cx2_4', 'hpc_c_d_4', 'hpc_c_s_4',
                       'EMG1_4', 'EMG2_4', 'cb_deep_4', 'cb_sup_4', 'hp_c_d_4', 'hpc_c_s_4',

                       'Opto']
        channel_types=['eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'stim']



    if headstage_number == 3:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,48]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

        datatp=data.transpose()#Array from openephys has to be transposed to match RawArray MNE function to create.
        del data
        del data_adc

        'Below I make the channel names and channel types, this should go in the parameteres file later.'

        n_channels=49

        channel_names=['hpc_mid_deep', 'hpc_mid_sup', 'hpc_ros_deep', 'hpc_ros_sup', 'pfc_deep',
                       'pfc_sup', 'cx1', 'cx2', 'hpc_contra_deep', 'hpc_contra_sup',
                       'EMG1', 'EMG2', 'cb_deep', 'cb_sup', 'hp_caudal_deep', 'hpc_caudal_sup',
                       'hpc_mid_d_2', 'hpc_mid_s_2', 'hpc_ros_d_2', 'hpc_ros_s_2', 'pfc_d_2',
                       'pfc_sup_2', 'cx1_2', 'cx2_2', 'hpc_ct_d_2', 'hpc_ct_s_2',
                       'EMG1_2', 'EMG2_2', 'cb_deep_2', 'cb_sup_2', 'hp_caud_d_2', 'hpc_caud_s_2',
                       'hpc_mid_d_3', 'hpc_mid_s_3', 'hpc_ros_d_3', 'hpc_ros_s_3', 'pfc_d_3',
                       'pfc_s_3', 'cx1_3', 'cx2_3', 'hpc_c_d_3', 'hpc_c_s_3',
                       'EMG1_3', 'EMG2_3', 'cb_deep_3', 'cb_sup_3', 'hp_c_d_3', 'hpc_c_s_3',
                       'Opto']
        channel_types=['eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'stim']

    if headstage_number ==2:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,32]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

        datatp=data.transpose()#Array from openephys has to be transposed to match RawArray MNE function to create.
        del data
        del data_adc

        n_channels=33

        channel_names=['hpc_mid_deep', 'hpc_mid_sup', 'hpc_ros_deep', 'hpc_ros_sup', 'pfc_deep',
                       'pfc_sup', 'cx1', 'cx2', 'hpc_contra_deep', 'hpc_contra_sup',
                       'EMG1', 'EMG2', 'cb_deep', 'cb_sup', 'hp_caudal_deep', 'hpc_caudal_sup',
                       'hpc_mid_d_2', 'hpc_mid_s_2', 'hpc_ros_d_2', 'hpc_ros_s_2', 'pfc_d_2',
                       'pfc_sup_2', 'cx1_2', 'cx2_2', 'hpc_ct_d_2', 'hpc_ct_s_2',
                       'EMG1_2', 'EMG2_2', 'cb_deep_2', 'cb_sup_2', 'hp_caud_d_2', 'hpc_caud_s_2',
                       'Opto']
        channel_types=['eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'stim']

    if headstage_number ==1:

        'Add Opto Stim Channel to Data'

        data= np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
        data[:,16]=(data_adc[:,0]*300) #Multiply by 300 to have about the same scale for optogenetics.

        datatp=data.transpose()#Array from openephys has to be transposed to match RawArray MNE function to create.
        del data
        del data_adc

        n_channels=17

        channel_names=['hpc_mid_deep', 'hpc_mid_sup', 'hpc_ros_deep', 'hpc_ros_sup', 'pfc_deep',
                       'pfc_sup', 'cx1', 'cx2', 'hpc_contra_deep', 'hpc_contra_sup',
                       'EMG1', 'EMG2', 'cb_deep', 'cb_sup', 'hp_caudal_deep', 'hpc_caudal_sup',
                       'Opto']
        channel_types=['eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg','emg','emg','eeg','eeg','eeg','eeg',
                       'stim']


    'This creates the info that goes with the channels, which is names, sampling rate, and channel types.'
    info = mne.create_info(channel_names, prm.get_sampling_rate(), channel_types)


    'This makes the object that contains all the data and info about the channels.'
    'Computations like plotting, averaging, power spectrums can be performed on this object'

    custom_raw = mne.io.RawArray( datatp, info)
    del datatp
    return custom_raw


def load_32_EEG_downsampled_bystate(foldername, montage_name, source_number, brain_states, downsampling, amp_filter = 750):
    """
    Load open ephys 32 electrode data to a numpy array
    after downsampling it. 
    then it creates numpy arrays for the brain states

    foldername: folder where the recordings are
    montage_name: name of the montage that contains the electrode coordinates
    source number: prefix of the electrode recordings files
    downsampling: integer with the number the sampling is going to be reduced by
    amp_filter: threshold amplitude to eliminate a sample
    
    """

    'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
    data=loadFolderToArray(foldername, channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = source_number)
    data_aux=loadFolderToArray(foldername, channels = 'all', chprefix = 'AUX', dtype = float, session = '0', source = source_number)

    data = np.vstack([np.transpose(data), np.transpose(data_aux)])

    # Add the brain states without downsampling them (but yes repeating)
    # As the bins are for every 5 seconds, we need to create an array repeating those states per each ms*downsampling
    brain_states_ms = np.repeat(brain_states, 5000)
    
    # 36 rows array array. First brain states row, then 32 eeg, then 3 emg electrodes
    state_voltage_array = np.vstack([brain_states_ms, data[:, 0:np.size(brain_states_ms)]])

    # Amplitude filter. Delete samples in the same sample of every electrode if one sample in one electrode
    # gets over a threshold amplitude    
    for i in np.arange(32):
      state_voltage_array = state_voltage_array[:, abs(state_voltage_array[i+1,:]) < amp_filter]
    
    # split the recording into the different brain states
    volt_wake = state_voltage_array[1:, state_voltage_array[0,:] == 0] # :1 because we do not want the brain_states raw anymore
    volt_NoREM = state_voltage_array[1:, state_voltage_array[0,:] == 1]
    volt_REM = state_voltage_array[1:, state_voltage_array[0,:] == 2]
    volt_convuls = state_voltage_array[1:, state_voltage_array[0,:] == 4]    

    # Doing the downsampling now, the decimate function filtering will smooth the edging we have
    # produced in both the amplitude filtering and the split by brain state
    
    # It will return the filtered and decimated whole thing, but without the brain states row -> [1:, :]
    raw_data_array = decimate(state_voltage_array[1:, :], downsampling, axis = 1)

    # And the decimated/downsampled (and filtered) arrays for every brain state
    volt_wake = decimate(volt_wake, downsampling, axis = 1)
    volt_NoREM = decimate(volt_NoREM, downsampling, axis = 1)
    volt_REM = decimate(volt_REM, downsampling, axis = 1)
    volt_convuls = decimate(volt_convuls, downsampling, axis = 1)

    # deleted big arrays of unused data
    del data
    del data_aux
    del brain_states_ms
    del state_voltage_array   
    
    return raw_data_array, volt_wake, volt_NoREM, volt_REM, volt_convuls

def load_32_EEG_downsampled_sleep(foldername, montage_name, source_number, brain_states, downsampling, amp_filter = 750):
    """
    Load open ephys 32 electrode sleep data to a numpy array
    after downsampling it. 
    then it creates numpy arrays for the brain states

    foldername: folder where the recordings are
    montage_name: name of the montage that contains the electrode coordinates
    source number: prefix of the electrode recordings files
    downsampling: integer with the number the sampling is going to be reduced by
    amp_filter: threshold amplitude to eliminate a sample
    
    """

    'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
    data=loadFolderToArray(foldername, channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = source_number)
    data_aux=loadFolderToArray(foldername, channels = 'all', chprefix = 'AUX', dtype = float, session = '0', source = source_number)

    data = np.vstack([np.transpose(data), np.transpose(data_aux)])

    # Add the brain states without downsampling them (but yes repeating)
    # As the bins are for every 5 seconds, we need to create an array repeating those states per each ms*downsampling
    brain_states_ms = np.repeat(brain_states, 5000)
    
    # 36 rows array array. First brain states row, then 32 eeg, then 3 emg electrodes
    state_voltage_array = np.vstack([brain_states_ms, data[:, 0:np.size(brain_states_ms)]])

    # split the recording into the different brain states
    volt_NoREM = state_voltage_array[1:, state_voltage_array[0,:] == 1]
    volt_REM = state_voltage_array[1:, state_voltage_array[0,:] == 2]    

    # Amplitude filter. Delete samples in the same sample of every electrode if one sample in one electrode
    # gets over a threshold amplitude    
    for i in np.arange(32):
      volt_NoREM = volt_NoREM[:, abs(volt_NoREM[i,:]) < amp_filter]
      volt_REM = volt_REM[:, abs(volt_REM[i,:]) < amp_filter]
    
    # And the decimated/downsampled (and filtered) arrays for every brain state
    volt_NoREM = decimate(volt_NoREM, downsampling, axis = 1)
    volt_REM = decimate(volt_REM, downsampling, axis = 1)
    
    # deleted big arrays of unused data
    del data
    del data_aux
    del brain_states_ms
    del state_voltage_array   
    
    return volt_NoREM, volt_REM


def load_32_EEG_downsampled_wake_conv(foldername, montage_name, source_number, brain_states, downsampling, amp_filter = 750):
    """
    Load open only wake and convulsion data 
    from ephys 32 electrode data to a numpy array
    after downsampling it. 
    then it creates numpy arrays for the brain states

    foldername: folder where the recordings are
    montage_name: name of the montage that contains the electrode coordinates
    source number: prefix of the electrode recordings files
    downsampling: integer with the number the sampling is going to be reduced by
    amp_filter: threshold amplitude to eliminate a sample
    
    """

    'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
    data=loadFolderToArray(foldername, channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = source_number)
    data_aux=loadFolderToArray(foldername, channels = 'all', chprefix = 'AUX', dtype = float, session = '0', source = source_number)

    data = np.vstack([np.transpose(data), np.transpose(data_aux)])

    # Add the brain states without downsampling them (but yes repeating)
    # As the bins are for every 5 seconds, we need to create an array repeating those states per each ms*downsampling
    brain_states_ms = np.repeat(brain_states, 5000)
    
    # 36 rows array array. First brain states row, then 32 eeg, then 3 emg electrodes
    state_voltage_array = np.vstack([brain_states_ms, data[:, 0:np.size(brain_states_ms)]])

    # split the recording into the different brain states
    volt_middle = state_voltage_array[:, state_voltage_array[0,:] != 1] # :1 because we do not want the brain_states raw anymore
    state_voltage_array = volt_middle[:, volt_middle[0,:] != 2]
    
    # Amplitude filter. Delete samples in the same sample of every electrode if one sample in one electrode
    # gets over a threshold amplitude    
    for i in np.arange(32):
      state_voltage_array = state_voltage_array[:, abs(state_voltage_array[i+1,:]) < amp_filter]

    # It will return the filtered and decimated whole thing, but without the brain states row -> [1:, :]
    raw_data_array = decimate(state_voltage_array[1:, :], downsampling, axis = 1)

    # deleted big arrays of unused data
    del data
    del data_aux
    del brain_states_ms
    del state_voltage_array   
    
    return raw_data_array

'The function below loads individual 32 channel probe recordings'
def load_32_EEG(foldername, montage_name, source_number):

    'Below are 2 functions from OpenEphys to load data channels and auxilary (accelerometer) channels'
    data=loadFolderToArray(foldername, channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = source_number)
    data_aux=loadFolderToArray(foldername, channels = 'all', chprefix = 'AUX', dtype = float, session = '0', source = source_number)

    #data=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'CH', dtype = float, session = '0', source = '101')
    #data_aux=loadFolderToArray(prm.get_filepath(), channels = 'all', chprefix = 'AUX', dtype = float, session = '0', source = '101')

    'Below we append a line to the data array and add the accelrometer data. We transpose to fit the MNE data format.'
    data = np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
    data[:,32]=data_aux[:,0]*800

    #To work with standard MNE montages it is necessary to add 3 extra channels for the position of left and right ear and nose.
    data = np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
    data[:,33]=data[:,32]
    data = np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
    data[:,34]=data[:,32]
    data = np.append(data, (np.zeros((data.shape[0],1), dtype=int64)), axis=1)
    data[:,35]=data[:,32]

    datatp=data.transpose() #Array from openephys has to be transposed to match RawArray MNE function to create.
    del data

    if isinstance('montage_name', str):
        montage = mne.channels.read_montage(montage_name)
    else:
        print("The montage name is not valid")

    channel_types=['eeg','eeg','eeg','eeg','eeg','eeg', 'eeg', 'eeg',
                   'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg'
                   ,'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg'
                   ,'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg', 'emg', 'emg', 'emg', 'emg']

    info = mne.create_info(montage.ch_names, prm.get_sampling_rate(), ch_types=channel_types,
                           montage=montage)

    'This makes the object that contains all the data and info about the channels.'
    'Computations like plotting, averaging, power spectrums can be performed on this object'

    custom_raw = mne.io.RawArray( datatp, info)
    del datatp
    return custom_raw

def parse_dat(fn, number_of_channels = 16, sample_rate = 1000):
  '''Load a .dat file by interpreting it as int16 and then de-interlacing the 16 channels'''
  sample_datatype = 'int16'
  display_decimation = 1

  # Load the raw (1-D) data
  dat_raw = np.fromfile(fn, dtype=sample_datatype)

  # Reshape the (2-D) per channel data
  step = number_of_channels * display_decimation
  dat_chans = [dat_raw[c::step] for c in range(number_of_channels)]

  # Build the time array
  t = np.arange(len(dat_chans[0]), dtype=float) / sample_rate

  return dat_chans, t

def load_16_EEG_taini_down_by_state(file_route, brain_states, downsampling, amp_filter, first_sample, sample_rate):
  n_channels = 16
  
  dat_chans, t=parse_dat(file_route, n_channels, sample_rate)   
  data=np.array(dat_chans)

  # the emg electrodes are in position 1 & 14, put them at the end. 
  # create extra emgchannels to follow mne montage requirement
  (original_elect, n_samples) = data.shape
  final_data = np.zeros((16, n_samples))
  final_data[0:14, :] = data[[0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15],:]
  final_data[14:16, :] = data[[1,14],:]
  
  del(dat_chans)
  del(data)

  # Add the brain states without downsampling them (but yes repeating)
  # As the bins are for every 5 seconds, we need to create an array repeating those states per each ms*downsampling
  brain_states_ms = np.repeat(brain_states, 5*sample_rate)  
  
  final_data = final_data[:, first_sample:first_sample+np.size(brain_states_ms)]

  # 17 rows array. First brain states row, then 14 eeg, then 2 emg electrodes
  state_voltage_array = np.vstack([brain_states_ms, final_data])
  del final_data

  # Amplitude filter. Delete samples in the same sample of every electrode if one sample in one electrode
  # gets over a threshold amplitude    
  for i in np.arange(14):
    state_voltage_array = state_voltage_array[:, abs(state_voltage_array[i+1,:]) < amp_filter]
  
  # split the recording into the different brain states
  volt_wake = state_voltage_array[1:, state_voltage_array[0,:] == 0] # :1 because we do not want the brain_states raw anymore
  volt_NoREM = state_voltage_array[1:, state_voltage_array[0,:] == 1]
  volt_REM = state_voltage_array[1:, state_voltage_array[0,:] == 2]
  volt_convuls = state_voltage_array[1:, state_voltage_array[0,:] == 4]    

  # Doing the downsampling now, the decimate function filtering will smooth the edging we have
  # produced in both the amplitude filtering and the split by brain state
  
  if downsampling == 1:
    return state_voltage_array, volt_wake, volt_NoREM, volt_REM, volt_convuls
  else:
    # It will return the filtered and decimated whole thing, but without the brain states row -> [1:, :]
    raw_data_array = decimate(state_voltage_array[1:, :], downsampling, axis = 1).astype(int16)
    del state_voltage_array

    # And the decimated/downsampled (and filtered) arrays for every brain state
    volt_wake = decimate(volt_wake, downsampling, axis = 1).astype(int16)
    volt_NoREM = decimate(volt_NoREM, downsampling, axis = 1).astype(int16)
    volt_REM = decimate(volt_REM, downsampling, axis = 1).astype(int16)
    if volt_convuls.size > 27:
        volt_convuls = decimate(volt_convuls, downsampling, axis = 1).astype(int16)
    else: 
        print(' Not enough convulsion time for downsampling')

    return raw_data_array, volt_wake, volt_NoREM, volt_REM, volt_convuls


def load_16_EEG_taini_down(file_route, downsampling, amp_filter, first_sample, sample_rate):
  n_channels = 16
  
  dat_chans, t=parse_dat(file_route, n_channels, sample_rate)   
  data=np.array(dat_chans, dtype=int16)

  # the emg electrodes are in position 1 & 14, put them at the end. 
  # create extra emgchannels to follow mne montage requirement
  (original_elect, n_samples) = data.shape
  final_data = np.zeros((16, n_samples), dtype=int16)
  final_data[0:14, :] = data[[0, 2, 3, 4, 5, 6, 7, 8, 9, 10, 11, 12, 13, 15],:]
  final_data[14:16, :] = data[[1,14],:]
  
  del(dat_chans)
  del(data)

  # Amplitude filter. Delete samples in the same sample of every electrode if one sample in one electrode
  # gets over a threshold amplitude    
  for i in np.arange(14):
    final_data = final_data[:, abs(final_data[i+1,:]) < amp_filter]
  
  # Doing the downsampling now, the decimate function filtering will smooth the edging we have
  # produced in both the amplitude filtering and the split by brain state
  
  # It will return the filtered and decimated whole thing
  if downsampling == 1:
    return final_data
  else:
    raw_data_array = decimate(final_data, downsampling, axis = 1).astype(int16)
    del final_data
    return raw_data_array

def npy32mne(filename, montage_name, sampling_rate):
    """
    Load numpy array file of 32 electrodes + 3 aux
    and converts it to mne format

    filename: route to the .npy file
    montage_name: route to the mne montage file
    
    """
    
    voltage_array = np.load(filename) # load
    
    if isinstance('montage_name', str):
        montage = mne.channels.read_custom_montage(montage_name)
    else:
        print("The montage name is not valid")

    channel_types=['eeg','eeg','eeg','eeg','eeg','eeg', 'eeg', 'eeg',
                   'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg'
                   ,'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg'
                   ,'eeg','eeg','eeg','eeg','eeg','eeg','eeg','eeg', 'emg', 'emg', 'emg']

    info = mne.create_info(montage.ch_names, sampling_rate, ch_types=channel_types)

    'This makes the object that contains all the data and info about the channels.'
    'Computations like plotting, averaging, power spectrums can be performed on this object'

    custom_raw = mne.io.RawArray(voltage_array, info)
    del voltage_array
    return custom_raw


def taininumpy2mne(npy_file, montage_name, sample_rate):
  ''' converts a .npy files containing 16 eeg electrodes data
      into mne format. 
      npy_data: location of the .npy file
      montage_name: location of the montage file
      sample_rate: sampling rate of the recording '''

  voltage_array = np.load(npy_file) 

  if isinstance('montage_name', str):
        montage = mne.channels.read_custom_montage(montage_name)
  else:
        print("The montage name is not valid")

  
  # 14 eeg channels, 2 emg 
  channel_types=['eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg',
                    'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'emg', 'emg']
  
  'This creates the info that goes with the channels, which is names, sampling rate and channel types.'
  info = mne.create_info(montage.ch_names, sample_rate, ch_types=channel_types)
  
  custom_raw = mne.io.RawArray(voltage_array, info)
  
  return custom_raw  

def tetrodesnumpy2mne(npy_file, montage_name, sample_rate):
  ''' converts a .npy files containing 16 eeg electrodes data
      from tetrodes into mne format. 
      npy_data: location of the .npy file
      montage_name: location of the montage file
      sample_rate: sampling rate of the recording '''

  voltage_array = np.transpose(np.load(npy_file))

  if isinstance('montage_name', str):
        montage = mne.channels.read_custom_montage(montage_name)
  else:
        print("The montage name is not valid")

  
  # 16 channels, 4 tetrodes 
  channel_types=['eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg',
                    'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg']
  
  'This creates the info that goes with the channels, which is names, sampling rate and channel types.'
  info = mne.create_info(montage.ch_names, sample_rate, ch_types=channel_types)
  
  custom_raw = mne.io.RawArray(voltage_array, info)
  
  return custom_raw  

def taininumpy2mnechannels(npy_file, montage_name, sample_rate, channels_list):
  ''' converts a .npy files containing 16 eeg electrodes data
      into mne format. 
      npy_data: location of the .npy file
      montage_name: location of the montage file
      sample_rate: sampling rate of the recording 
      channels_list: list with the number positions of the channels we want '''

  voltage_array = np.load(npy_file) 

  # New voltage array with just the number of channels we want
  array_size = np.shape(voltage_array)
  shorter_array = np.zeros((len(channels_list), array_size[1]))

  # 14 eeg channels, 2 emg to start with
  channel_types=['eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg',
                    'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'eeg', 'emg', 'emg']

  shorter_channel_types = []
  # Filling the shorter voltage array and the shorter channel list
  for i, channel in enumerate(channels_list):
    shorter_array[i] = voltage_array[channel]
    shorter_channel_types.append(channel_types[channel])

  # the montage will need to have only the channels passed in channel_list
  if isinstance('montage_name', str):
        montage = mne.channels.read_custom_montage(montage_name)
  else:
        print("The montage name is not valid")

  
  'This creates the info that goes with the channels, which is names, sampling rate and channel types.'
  info = mne.create_info(montage.ch_names, sample_rate, ch_types=shorter_channel_types)
  
  custom_raw = mne.io.RawArray(shorter_array, info)
  
  return custom_raw  


def electrode_combinations(montage_name, neighbors_dist, long_distance, recording, n_elect = 32):
  montage = mne.channels.read_custom_montage(montage_name)
  electrode_names = montage.ch_names[0:n_elect]
  
  electrode_pos = np.zeros((n_elect,2))
  for i in np.arange(n_elect):
    if recording == 'taini':
        electrode_pos[i] = montage.dig[i].get('r')[0:2].round(3)*50 # in mm
    else:
        electrode_pos[i] = montage.dig[i+3].get('r')[0:2].round(3)*50 # in mm

  distances_btw_electrodes = spatial.distance.pdist(electrode_pos, 'euclidean')

  nums = np.linspace(0, n_elect-1, n_elect, dtype = int)

  comb = combinations(nums, 2)
  # working with the combination element is difficult and it can only be assigned once -> it is transformed into a list
  comb_long_distance = list(comb)

  comb = combinations(nums, 2)
  comb_short_distance = list(comb)

  # indexes for the elements to delete
  indexes_to_delete_in_short_distance = []
  indexes_to_delete_in_long_distance = []

  nei = 0
  s_d = 0
  l_d = 0

  for i in range(len(distances_btw_electrodes)):
      if distances_btw_electrodes[i] <= neighbors_dist:
          indexes_to_delete_in_long_distance.append(i)
          indexes_to_delete_in_short_distance.append(i)
          nei += 1
      elif distances_btw_electrodes[i] <long_distance:
          indexes_to_delete_in_long_distance.append(i)
          s_d += 1
      else:
          indexes_to_delete_in_short_distance.append(i)
          l_d += 1

  long_dist_electrodes = np.delete(distances_btw_electrodes, indexes_to_delete_in_long_distance)
  short_dist_electrodes = np.delete(distances_btw_electrodes, indexes_to_delete_in_short_distance)

  # when it deletes an element, it is necessary to update the indexes substracting one to the total of them the list has.
  indexes_already_del = 0
  for i in range(len(indexes_to_delete_in_long_distance)):
      del comb_long_distance[indexes_to_delete_in_long_distance[i] - indexes_already_del]
      indexes_already_del += 1

  indexes_already_del = 0
  for i in range(len(indexes_to_delete_in_short_distance)):
      del comb_short_distance[indexes_to_delete_in_short_distance[i] - indexes_already_del]
      indexes_already_del += 1

  return comb_short_distance, comb_long_distance


def create_epochs(analysis_times, sampling_rate): #Makes epoch file for MNE of stimulation times.

    num_rows, num_cols=analysis_times.shape
    epochs= tile(0, (num_rows, 3))
    for n in range(0, num_rows):
        start_time=(analysis_times.item(n,0))
        epochs[n][0]=start_time*sampling_rate
        epochs[n][2]=n


    return epochs

def create_brain_state_epochs(analysis_times, sampling_rate): #Makes epoch file for MNE of stimulation times.

    num_rows, num_cols=analysis_times.shape
    epochs= tile(0, (num_rows, 3))
    for n in range(0, num_rows):
        start_time=(analysis_times.item(n,1))
        ident=(analysis_times.item(n,0))
        epochs[n][0]=start_time
        epochs[n][2]=ident
        epochs[n][1]=n


    return epochs

# Read the brain states of a recording from a .xls file
# Every cell is the state during a bin of 10 s
# Make sure values are not empty.
def import_brain_states(excelfile):
    print(" Openning Excelfile " + excelfile)
    book= xlrd.open_workbook(excelfile)
    sheet_names = book.sheet_names()
    sheet= book.sheet_by_name(sheet_names[0])
    brain_state = np.zeros(shape=(sheet.nrows, 1))

    for n in range(0, sheet.nrows-1):
        cell1=sheet.cell(n, 0)
        cell1_value=cell1.value
        brain_state[n, 0]=cell1_value

    return brain_state

def actual_stim_times(data, sampling_rate):##for use with normal opto

    times=[]
    times=data[:,64]>300
    start_times=[]

    for n in range(len(times)):
        if times[n] == True:
            start_times.append(n)

    stim_times=[]

    for n in range(len(start_times)):
        x=n-1
        if n ==0:
            stim_times.append(start_times[n]/sampling_rate)
        elif start_times[n]-start_times[n-1]>(sampling_rate*20):
            stim_times.append(start_times[n]/sampling_rate)


    stimulations= asarray(stim_times)

    return stimulations





def import_brain_state(excelfile): #Import analysis times for optogenetic on and control times from excel.
    #Make sure values are not empty.
    #Save file as .xls"
    print(" Openning Brain State Data " + excelfile)
    book= xlrd.open_workbook(excelfile)
    sheet_names = book.sheet_names()
    sheet= book.sheet_by_name(sheet_names[0])
    analysis_times= zeros(shape=(sheet.nrows, 3))
    print(analysis_times.shape)
    for n in range(0, sheet.nrows):
        cell1=sheet.cell(n, 0)
        cell1_value=cell1.value
        analysis_times[n, 0]=cell1_value

        cell2=sheet.cell(n, 1)
        cell2_value=cell2.value
        analysis_times[n, 1]=cell2_value

        cell3=sheet.cell(n, 2)
        cell3_value=cell3.value
        analysis_times[n, 2]=cell3_value

    return analysis_times


def import_channel_combo(excelfile):
    print(" Openning Channel Data " + excelfile)
    book= xlrd.open_workbook(excelfile)
    sheet_names = book.sheet_names()
    sheet= book.sheet_by_name(sheet_names[0])
    channel_combo= zeros(shape=(sheet.nrows, 2))

    for n in range(0, sheet.nrows):
        cell1=sheet.cell(n, 0)
        cell1_value=cell1.value
        channel_combo[n, 0]=cell1_value
        cell2=sheet.cell(n, 1)
        cell2_value=cell2.value
        channel_combo[n, 1]=cell2_value

    return channel_combo


def import_spreadsheet(excelfile): #Import analysis times for optogenetic on and control times from excel.
    #Make sure values are not empty.
    #Save file as .xls"
    print(" Openning Excelfile " + excelfile)
    book= xlrd.open_workbook(excelfile)
    sheet_names = book.sheet_names()
    sheet= book.sheet_by_name(sheet_names[0])
    analysis_times= zeros(shape=(sheet.nrows - 1, 5))

    for n in range(0, sheet.nrows-1):
        cell1=sheet.cell(n+1, 3)
        cell1_value=cell1.value
        analysis_times[n, 0]=cell1_value

        # Every cell is empty in the next columns in 300abs_thresh.xls!!
        # =============================================================================
        #         cell2=sheet.cell(n+1, 4)
        #         cell2_value=cell2.value
        #         analysis_times[n, 1]=cell2_value
        #         cell5=sheet.cell(n+1, 5)
        #         cell5_value=cell5.value
        #         analysis_times[n, 2]=cell5_value
        #         cell6=sheet.cell(n+1, 6)
        #         cell6_value=cell6.value
        #         analysis_times[n, 3]=cell6_value
        #         cell7=sheet.cell(n+1, 7)
        #         cell7_value=cell7.value
        #         analysis_times[n, 4]=cell7_value
        # =============================================================================

    return analysis_times


def sub_time_data(data, start_time, end_time, sampling_rate): #Gets time axis and data of specific times.

    prm.set_filelength(len(data))
    filelength=prm.get_filelength()
    timelength=filelength/sampling_rate
    time_axis = linspace(start_time, end_time, ((end_time-start_time)*sampling_rate))

    index_start = start_time*sampling_rate
    index_end = end_time*sampling_rate
    sub_data = data[int(index_start):int(index_end)]

    return time_axis, sub_data

# https://stackoverflow.com/a/2415343/12094741
def weighted_avg_sem(values, weights, id_rec):
    """
    Return the weighted average and standard error of the mean.

    values, weights -- Numpy ndarrays with the same shape.
    """
    average = np.average(values, weights=weights)
    # Fast and numerically precise:
    variance = np.average((values-average)**2, weights=weights)
    if (variance < 0):
      print(f"For this recording: {id_rec}")
      print(f"variance was {variance}")
      print("values")
      print(np.array(values))
      print('weights')
      print(np.array(weights))
      print(f"average = {average}")
      variance = abs(np.average((values-average)**2, weights=weights))

    std_dev = math.sqrt(variance)
    sem_w = std_dev/len(values)
    return (average, sem_w)


def plot_all(data, sampling_rate, color):  #This allows for an initial plot of all the ddata.

    timemax=len(data)
    timelength = timemax/sampling_rate

    timeforplot = linspace(0, timelength, timemax)
    plt.plot(timeforplot, data, color)

    return


# Printing and exporting to excel individual coherences
def df_to_excel(filename, data_frame, sheet_n):
  """
    Write a dataframe in a particular sheet of a chosen excel file.

    If the file does not exist, it is created
    If the file already exists, it is not overwritten

    """
  file = pathlib.Path(filename)
  if file.exists ():
    book = load_workbook(filename)
    # xlsxwrite does not allow to open an excel file -> openpyxl
    writer = pd.ExcelWriter(filename, engine = 'openpyxl')
    writer.book = book
  else:
    writer = pd.ExcelWriter(filename, engine='xlsxwriter')

  data_frame.to_excel(writer, sheet_name=sheet_n)
  writer.save()
  writer.close()

# Calculates the closest below power of 2 to a target. Useful fur n_fft in Power spectrum
def power_of_two(target):
    x = 2
    change = 0
    power = 0
    for i in range(int(target)+1):
        number = x ** change
        change = i
        if number >= int(target):  # you had indentation errors here and following
            power = int(number/2)
            return power

